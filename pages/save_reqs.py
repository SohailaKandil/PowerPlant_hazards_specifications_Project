import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

try:
    uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"], label_visibility="collapsed")
except:
    st.warning("⚠️ There was an error uploading your file please try again later")
if uploaded_file and "selected_data" in st.session_state:
    try:
        # Load the workbook from uploaded file
        in_memory_file = BytesIO(uploaded_file.read())
        wb = load_workbook(in_memory_file)
        ws = wb.active  # Or use wb['SheetName'] if needed

        # Find first empty row starting from row 2
        start_row = 2
        for row in range(start_row, ws.max_row + 1):
            if ws.cell(row=row, column=1).value is None and ws.cell(row=row, column=2).value is None:
                next_empty_row = row
                break
        else:
            next_empty_row = ws.max_row + 1

        # Insert data
        for item, desc in zip(st.session_state["selected_data"], st.session_state["selected_data_desc"]):
            ws.cell(row=next_empty_row, column=1, value=item)
            ws.cell(row=next_empty_row, column=2, value=desc)
            next_empty_row += 1

        # Save to memory for download
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        st.success("✅ Data selections saved successfully.")

        # Download button
        st.download_button(
            label="📥 Download Updated Excel File",
            data=output,
            file_name="updated_output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"❌ Error: {e}")

elif uploaded_file is None:
    st.info("📤 Please upload your Excel file first.")
else:
    st.warning("⚠️ Please go to the Equipment details page first.")